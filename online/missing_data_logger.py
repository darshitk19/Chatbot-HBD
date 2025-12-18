from openpyxl import Workbook, load_workbook
from datetime import datetime
import json
import os

FILE_NAME = "missing_searches.xlsx"

def log_missing_query(query, results=None):
    if not os.path.exists(FILE_NAME):
        wb = Workbook()
        ws = wb.active
        ws.append(["query", "timestamp", "snapshot"])
        wb.save(FILE_NAME)

    wb = load_workbook(FILE_NAME)
    ws = wb.active

    ws.append([
        query,
        datetime.utcnow().isoformat(),
        json.dumps(results[:3]) if results else None
    ])

    wb.save(FILE_NAME)
